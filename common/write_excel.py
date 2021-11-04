import os
import sys
import time

import openpyxl
import xlrd
from openpyxl import load_workbook
from self import self
import easygui
from win32com.client import Dispatch
from openpyxl.styles import Font
from openpyxl.styles.colors import (BLACK, RED, GREEN)

from openpyxl import load_workbook
from collections import namedtuple



class mywrite:
    def write_cell(self, sheet_name, row, column, value, color='black'):
        """write cell value with color"""
        if isinstance(row, int) and isinstance(column, int):
            try:
                cell_obj = self.wb[sheet_name].cell(row, column)
                cell_obj.font = Font(color=self.RGBDict[color], bold=True)
                cell_obj.value = value
                self.wb.save(self.filename)
            except Exception as e:
                raise e
        else:
            raise TypeError('row and column must be type int')

    def write(self, myexcel, myshell, rowNum, colNum, results, color='black'):
        # wb = load_workbook("D:/testdata/gitami/pythonami/pythontest/data/接口自动化测试.xlsx")
        # wb2 = wb['测试用例']
        # wb2.cell(2, 13, '333')
        # wb.save("D:/testdata/gitami/pythonami/pythontest/data/接口自动化测试.xlsx")  # 保存
        wb = openpyxl.load_workbook(myexcel)
        # wb = xlrd.open_workbook(myexcel)
        page = len(wb.sheetnames)
        # page = len(wb.sheets())
        print("sda—："+str(myshell))
        wb2 = wb[myshell]

        ss = wb2.cell(rowNum, colNum, results)
        RGBDict = {'red': RED, 'green': GREEN, 'black': BLACK}
        ss.font = Font(color=RGBDict[color], bold=True)
        wb.save(myexcel)  # 保存
        wb.close()  #关闭
        print("成功")

    def check_excel_is_open(self) -> int:
        # file_path= "C:\\ProgramData\\xxx\\xxxx\\BmEquipmentLibrary"
        file_path = "C:\\ProgramData\\xxx\\xxxx"
        filenames = os.listdir(file_path)
        try:
            for filename in filenames:
                os.rename(file_path + "\\" + filename, file_path + "\\tempfile.xls")
                os.rename(file_path + "\\tempfile.xls", file_path + "\\" + filename)

        except OSError:
            easygui.msgbox('excel打开状态，请关闭后尝试卸载！')
            return sys.exit(2)
        return sys.exit(1001)

# 传入路径判断是否打开,如果打开就关闭
def fileisopen(filepath):
    xlApp = Dispatch('Excel.Application')

    if xlApp.workbooks.count == 0:
        return False
    else:
        print(xlApp.workbooks.count)
        for i in range(1,xlApp.workbooks.count+1):
            realpth = xlApp.workbooks(i).Path+"\\"+xlApp.workbooks(i).Name
            print(i)
            print(realpth)
            #转化格式
            realpths=realpth.replace("\\","/")
            print(realpths.lower())
            print(filepath.lower())
            #lower()大写字母转化为小写字母
            if realpths.lower() == filepath.lower():
                xlApp.Workbooks(i).Close()  # 关闭当前打开的文件,不保存文件
                return True
        # xlApp.Quit()  #这个命令会把所有打开的excel文件关闭
        return False


if __name__ == '__main__':
    # file_path = "D:/testdata/gitami/pythonami/pythontest/data/接口自动化测试.xlsx"
    file_path = "D:/testdata/gitami/pythonami/superlucyjr/data/接口自动化测试.xlsx"
    #判断是否打开，打开就关闭
    fileopen = fileisopen(file_path)
    print(fileopen)
    # sheet_name = '测试用例'
    # ss = mywrite()
    # ss.write(myexcel=file_path, myshell=sheet_name, rowNum=2, colNum=15, results='passone')

